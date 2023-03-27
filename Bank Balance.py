import fitz
import os
from os import getcwd, chdir
from glob import glob as glob
import time
import openpyxl as xl
import sys

start_time = time.time()


def locate(values, worksheets):
    position_info = []
    for row_s in worksheets.iter_rows(min_col=1, max_col=worksheets.max_column):
        for cell_s in row_s:
            if cell_s.value == values:
                position_info.extend([cell_s.row, cell_s.column, cell_s.coordinate])
    if not position_info:
        pass
    else:
        return position_info


need_folders = ['Altura', 'CBT', 'First Foundation', 'LPL', 'Old']
sub_folders = [name for name in os.listdir(getcwd()) if os.path.isdir(os.path.join(getcwd(), name))]
diff = set(sub_folders).difference(set(need_folders))

if bool(diff != set()):
    print(f"Please rename or remove these folders: {list(diff)}")
    sys.exit()
elif bool(diff == set()):
    pass

sept_path = fr"C:\Users\{os.getlogin()}\PycharmProjects\pythonProject\Programs\Bank Balance Test\Sept Properties Invst - Template.xlsm"
wb = xl.load_workbook(sept_path, keep_vba=True)
ws = wb.worksheets[0]

dec_path = fr"C:\Users\{os.getlogin()}\PycharmProjects\pythonProject\Programs\Bank Balance Test\Dec Properties Invst - Template.xlsm"
wb1 = xl.load_workbook(dec_path, keep_vba=True)
ws1 = wb1.worksheets[0]

sept_acc_num = []
dec_acc_num = []

for i in range(6, ws.max_row):
    sept_acc_num.append(ws.cell(row=i, column=3).value)
for i in range(6, ws1.max_row):
    dec_acc_num.append(ws1.cell(row=i, column=4).value)

sept_acc_num_list = [str(x) for x in list(filter(None, sept_acc_num))]
dec_acc_num_list = [str(x) for x in list(filter(None, dec_acc_num))]

four_sept_acc_num_list = [str(x)[-4:] for x in list(filter(None, sept_acc_num))]
four_dec_acc_num_list = [str(x)[-4:] for x in list(filter(None, dec_acc_num))]


altura_dec_acc_num_list = [str(x)[-7:] for x in list(filter(None, dec_acc_num))]
altura_dec_list = []

for i in altura_dec_acc_num_list:
    if '-' in i:
        altura_dec_list.append(i)


cbt_directory = 'CBT'
curr_dir = getcwd()
chdir(cbt_directory)

cbt_list = glob('*.pdf')

cbt_sept_tup_list = []
cbt_dec_tup_list = []

for pdf in cbt_list:
    with fitz.open(pdf) as pdf_obj:
        pagecount = pdf_obj.page_count
        all_pages_text = []
        for i in range(0, pagecount):
            page_text = pdf_obj[i].get_text("words", sort=False)
            all_pages_text += page_text
        for i in sept_acc_num_list:
            for idx, j in enumerate(all_pages_text):
                list_list = []
                if i == j[4] and '$' in all_pages_text[idx + 1][4]:
                    list_list.append(int(i))
                    end_bal = all_pages_text[idx + 1][4][1:]
                    list_list.append(end_bal)
                    cbt_sept_tup_list.append((tuple(list_list)))

        for i in dec_acc_num_list:
            for idx, j in enumerate(all_pages_text):
                list_list = []
                if i == j[4] and '$' in all_pages_text[idx + 1][4]:
                    list_list.append(int(i))
                    end_bal = all_pages_text[idx + 1][4][1:]
                    list_list.append(end_bal)
                    cbt_dec_tup_list.append((tuple(list_list)))


lpl_directory = 'LPL'
curr_dir = getcwd()
# this os.chdir to bring the directory up a folder, since the previous one brought the folder down a level.
os.chdir('..')
chdir(lpl_directory)

lpl_list = glob('*.pdf')

lpl_sept_tup_list = []
lpl_dec_tup_list = []

for pdf in lpl_list:
    with fitz.open(pdf) as pdf_obj:
        pagecount = pdf_obj.page_count
        all_pages_text = []
        for i in range(0, pagecount):
            page_text = pdf_obj[i].get_text("words", sort=False)
            all_pages_text += page_text


        for i in sept_acc_num_list:
            for idx, j in enumerate(all_pages_text):
                list_list = []
                if i == j[4] and all_pages_text[idx - 1][4] == 'Number:':
                    for idx1, k in enumerate(all_pages_text):
                        if k[4] == 'Invested':
                            list_list.append(i)
                            end_bal = all_pages_text[idx1 - 20][4][1:]
                            list_list.append(end_bal)
                            lpl_sept_tup_list.append((tuple(list_list)))
#
        for i in dec_acc_num_list:
            for idx, j in enumerate(all_pages_text):
                list_list = []
                if i == j[4] and all_pages_text[idx - 1][4] == 'Number:':
                    for idx1, k in enumerate(all_pages_text):
                        if k[4] == 'Invested':
                            list_list.append(i)
                            end_bal = all_pages_text[idx1 - 20][4][1:]
                            list_list.append(end_bal)
                            lpl_dec_tup_list.append((tuple(list_list)))


ff_directory = 'First Foundation'
curr_dir = getcwd()
# this os.chdir to bring the directory up a folder, since the previous one brought the folder down a level.
os.chdir('..')
chdir(ff_directory)

ff_list = glob('*.pdf')

ff_sept_tup_list = []
ff_dec_tup_list = []

for pdf in ff_list:
    with fitz.open(pdf) as pdf_obj:
        pagecount = pdf_obj.page_count
        all_pages_text = []
        for i in range(0, pagecount):
            page_text = pdf_obj[i].get_text("words", sort=False)
            all_pages_text += page_text
        for i in four_dec_acc_num_list:
            for idx, j in enumerate(all_pages_text):
                list_list = []
                if i == j[4][-4:] and all_pages_text[idx - 1][4] == 'Promo':
                    list_list.append(i)
                    end_bal = all_pages_text[idx + 1][4][1:]
                    list_list.append(end_bal)
                    ff_dec_tup_list.append((tuple(list_list)))


altura_directory = 'Altura'
curr_dir = getcwd()
# this os.chdir to bring the directory up a folder, since the previous one brought the folder down a level.
os.chdir('..')
chdir(altura_directory)

altura_list = glob('*.pdf')

altura_dec_tup_list = []

for pdf in altura_list:
    with fitz.open(pdf) as pdf_obj:
        pagecount = pdf_obj.page_count
        all_pages_text = []
        for i in range(0, pagecount):
            page_text = pdf_obj[i].get_text("words", sort=False)
            all_pages_text += page_text
        for i in altura_dec_list:
            for idx, j in enumerate(all_pages_text):
                list_list = []
                if i == j[4][-7:]:
                    list_list.append(i)
                    end_bal = all_pages_text[idx + 17][4][1:]
                    list_list.append(end_bal)
                    altura_dec_tup_list.append((tuple(list_list)))


c = locate("California", ws)[1]
c1 = locate("LPL", ws)[1]

for i in cbt_sept_tup_list:
    ws.cell(row=locate(i[0], ws)[0], column=c).value = float(i[1].replace(',', ''))
for i in lpl_sept_tup_list:
    ws.cell(row=locate(i[0], ws)[0], column=c1).value = float(i[1].replace(',', ''))



c = locate("California", ws1)[1]
c1 = locate("LPL", ws1)[1]
c2 = locate("FOUNDATION", ws1)[1]
c3 = locate("Altura", ws1)[1]

for i in cbt_dec_tup_list:
    ws1.cell(row=locate(i[0], ws1)[0], column=c).value = float(i[1].replace(',', ''))
for i in lpl_dec_tup_list:
    ws1.cell(row=locate(i[0], ws1)[0], column=c1).value = float(i[1].replace(',', ''))

# Since the first element of tuple only contains the last 4 account numbers, I have to compare with the original list of
# account numbers, so if the last 4 digits match with the last 4 digits of an element in the full account number list,
# I am able to use the full account number element, which is an integer in the Excel, to locate the column it's in.
for i in ff_dec_tup_list:
    for j in dec_acc_num_list:
        if i[0] == j[-4:]:
            ws1.cell(row=locate(int(j), ws1)[0], column=c2).value = float(i[1].replace(',', ''))

for i in altura_dec_tup_list:
    for j in dec_acc_num_list:
        if i[0] == j[-7:]:
            ws1.cell(row=locate(j, ws1)[0], column=c3).value = float(i[1].replace(',', ''))

wb.save(sept_path)
wb1.save(dec_path)
wb.close()
wb1.close()
print("------- Program finished running in %s seconds. -------" % round((time.time() - start_time), 2))
